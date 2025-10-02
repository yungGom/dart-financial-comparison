"""
Enhanced Excel Generator Module
개선된 Excel 생성 모듈
"""
import pandas as pd
from typing import Dict, List, Any, Optional
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelGenerator:
    """개선된 Excel 생성기"""

    def __init__(self):
        # 스타일 정의
        self.header_font = Font(bold=True, size=12)
        self.title_font = Font(bold=True, size=14)
        self.header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    def _safe_format_amount(self, amount: Any) -> str:
        """
        안전하게 금액을 포맷팅하는 헬퍼 함수
        타입에 관계없이 안전하게 숫자로 변환 후 포맷팅

        Args:
            amount: 금액 (str, int, float 등)

        Returns:
            포맷팅된 금액 문자열 또는 "-"
        """
        if amount is None or amount == '' or amount == 0:
            return "-"

        try:
            # 문자열인 경우 쉼표 제거 후 변환
            if isinstance(amount, str):
                amount = amount.replace(',', '').strip()
                if not amount or amount == '-':
                    return "-"

            # 숫자로 변환
            numeric_amount = float(amount)

            # 0인 경우
            if numeric_amount == 0:
                return "-"

            # 포맷팅
            return f"{numeric_amount:,.0f}"

        except (ValueError, TypeError, AttributeError):
            # 변환 실패 시 원본 문자열 반환
            return str(amount) if amount else "-"

    def create_comparison_excel(
        self,
        comparison_data: Dict[str, Any],
        selected_accounts: List[str],
        note_data: Optional[Dict[str, Any]] = None
    ) -> io.BytesIO:
        """
        개선된 비교 Excel 파일 생성

        Args:
            comparison_data: 비교 데이터
            selected_accounts: 선택된 계정과목 ID 리스트
            note_data: 주석 정보 (감사인, 회계기준 등)

        Returns:
            Excel 파일 (BytesIO)
        """
        output = io.BytesIO()
        wb = Workbook()

        # 기본 시트 제거
        wb.remove(wb.active)

        # 1. 요약 시트 생성 (재무비율 비교)
        self._create_summary_sheet(wb, comparison_data)

        # 2. 연도별 데이터 시트 생성
        years = self._extract_years(comparison_data)
        for year in sorted(years, reverse=True):
            self._create_year_sheet(wb, year, comparison_data, selected_accounts, note_data)

        # 3. 주석 정보 시트 (선택사항)
        if note_data:
            self._create_notes_sheet(wb, note_data)

        wb.save(output)
        output.seek(0)
        return output

    def _create_summary_sheet(self, wb: Workbook, comparison_data: Dict[str, Any]):
        """요약 시트 생성 - 재무비율 비교"""
        ws = wb.create_sheet("요약_재무비율")

        # 제목
        ws['A1'] = "재무비율 비교 분석"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:J1')

        # 헤더
        headers = [
            '기업명', '연도', '유동비율(%)', '부채비율(%)', '자기자본비율(%)',
            'ROE(%)', 'ROA(%)', '영업이익률(%)', '순이익률(%)', '총자산회전율'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        # 데이터 입력
        row_idx = 4
        for key, data in comparison_data.items():
            if 'ratios' in data:
                company_name = data.get('company_name', '')
                year = data.get('year', '')
                ratios = data['ratios']

                row_data = [
                    company_name,
                    year,
                    ratios.get('stability', {}).get('current_ratio', 0),
                    ratios.get('stability', {}).get('debt_to_equity', 0),
                    ratios.get('stability', {}).get('equity_ratio', 0),
                    ratios.get('profitability', {}).get('roe', 0),
                    ratios.get('profitability', {}).get('roa', 0),
                    ratios.get('profitability', {}).get('operating_margin', 0),
                    ratios.get('profitability', {}).get('net_margin', 0),
                    ratios.get('activity', {}).get('asset_turnover', 0)
                ]

                for col, value in enumerate(row_data, 1):
                    # 숫자 컬럼인 경우 안전한 포맷팅 적용
                    if col > 2 and value != 0:
                        try:
                            value = float(value) if value else 0
                        except (ValueError, TypeError):
                            value = 0

                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = self.border
                    if col > 2:  # 숫자 데이터는 중앙 정렬
                        cell.alignment = self.center_alignment

                row_idx += 1

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            # MergedCell 처리를 위한 예외 처리 추가
            try:
                if hasattr(column[0], 'column_letter'):
                    column_letter = column[0].column_letter
                else:
                    continue
            except:
                continue

            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if max_length > 0:
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width

    def _create_year_sheet(
        self,
        wb: Workbook,
        year: str,
        comparison_data: Dict[str, Any],
        selected_accounts: List[str],
        note_data: Optional[Dict[str, Any]]
    ):
        """연도별 상세 데이터 시트 생성"""
        ws = wb.create_sheet(f"{year}년")

        # 해당 연도의 기업들 추출
        year_companies = {}
        for key, data in comparison_data.items():
            if data.get('year') == year:
                company_name = data.get('company_name', '').strip()  # 공백 제거하여 일관성 유지
                year_companies[company_name] = data

        if not year_companies:
            return

        # 제목
        ws['A1'] = f"{year}년 재무정보 비교"
        ws['A1'].font = self.title_font
        ws.merge_cells(f'A1:{chr(65 + len(year_companies))}1')

        # 회사명 헤더
        companies = list(year_companies.keys())
        ws['A3'] = '항목'
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws['A3'].border = self.border

        for col, company in enumerate(companies, 2):
            cell = ws.cell(row=3, column=col, value=company)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

        row_idx = 4

        # 감사 정보 섹션
        if note_data:
            # 감사인
            ws.cell(row=row_idx, column=1, value="감사인").font = Font(bold=True)
            for col, company in enumerate(companies, 2):
                audit_info = note_data.get(company, {}).get(year, {})
                ws.cell(row=row_idx, column=col, value=audit_info.get('auditor', '-'))
            row_idx += 1

            # 회계기준
            ws.cell(row=row_idx, column=1, value="회계기준").font = Font(bold=True)
            for col, company in enumerate(companies, 2):
                audit_info = note_data.get(company, {}).get(year, {})
                ws.cell(row=row_idx, column=col, value=audit_info.get('accounting_standard', '-'))
            row_idx += 1

            # 감사의견
            ws.cell(row=row_idx, column=1, value="감사의견").font = Font(bold=True)
            for col, company in enumerate(companies, 2):
                audit_info = note_data.get(company, {}).get(year, {})
                ws.cell(row=row_idx, column=col, value=audit_info.get('audit_opinion', '-'))
            row_idx += 1

            # 감가상각방법
            ws.cell(row=row_idx, column=1, value="감가상각방법").font = Font(bold=True)
            for col, company in enumerate(companies, 2):
                audit_info = note_data.get(company, {}).get(year, {})
                depreciation = audit_info.get('depreciation_policy', {})
                method = depreciation.get('method', '-') if isinstance(depreciation, dict) else '-'
                ws.cell(row=row_idx, column=col, value=method)
            row_idx += 1

            # 내용연수 (유형자산별)
            # 먼저 모든 회사에서 사용하는 자산 유형 수집
            all_asset_types = set()
            for col, company in enumerate(companies, 2):
                audit_info = note_data.get(company, {}).get(year, {})
                useful_life = audit_info.get('useful_life', {})
                if isinstance(useful_life, dict):
                    all_asset_types.update(useful_life.keys())

            # 자산 유형별 내용연수 표시
            if all_asset_types:
                for asset_type in sorted(all_asset_types):
                    ws.cell(row=row_idx, column=1, value=f"내용연수({asset_type})").font = Font(bold=True)
                    for col, company in enumerate(companies, 2):
                        audit_info = note_data.get(company, {}).get(year, {})
                        useful_life = audit_info.get('useful_life', {})
                        life_value = useful_life.get(asset_type, '-') if isinstance(useful_life, dict) else '-'
                        ws.cell(row=row_idx, column=col, value=life_value)
                    row_idx += 1

            row_idx += 1  # 섹션 간 여백

        # 재무제표 데이터 섹션
        ws.cell(row=row_idx, column=1, value="[재무상태표]").font = Font(bold=True, color="000080")
        row_idx += 1

        # 계정과목별 데이터
        from .account_manager import account_manager
        account_mapping = account_manager.get_account_mapping()

        # 선택된 계정과목이 있으면 해당 계정만, 없으면 전체 계정 표시
        if selected_accounts:
            # 선택된 계정과목만 표시
            for account_id in selected_accounts:
                account_name = account_mapping.get(account_id, account_id)
                ws.cell(row=row_idx, column=1, value=account_name)

                for col, company in enumerate(companies, 2):
                    company_data = year_companies[company]
                    statements = company_data.get('statements', [])

                    # 해당 계정과목 찾기
                    value = "-"
                    unit = ""
                    for stmt in statements:
                        if stmt.get('account_id', '').strip() == account_id:
                            amount = stmt.get('thstrm_amount', 0)
                            # 금액 단위 확인
                            unit = account_manager.detect_amount_unit([stmt])
                            # 안전한 금액 포맷팅 사용
                            formatted_amount = self._safe_format_amount(amount)
                            if formatted_amount != "-":
                                value = formatted_amount
                                if unit and unit != "원":
                                    value += f" {unit}"
                            break

                    ws.cell(row=row_idx, column=col, value=value)

                row_idx += 1
        else:
            # 선택된 계정이 없으면 모든 계정 표시
            # 모든 회사의 계정과목 수집
            all_account_ids = set()
            for company_data in year_companies.values():
                for stmt in company_data.get('statements', []):
                    account_id = stmt.get('account_id', '').strip()
                    if account_id:
                        all_account_ids.add(account_id)

            # 계정과목별 데이터 표시
            for account_id in sorted(all_account_ids):
                account_name = account_mapping.get(account_id, account_id)
                ws.cell(row=row_idx, column=1, value=account_name)

                for col, company in enumerate(companies, 2):
                    company_data = year_companies[company]
                    statements = company_data.get('statements', [])

                    # 해당 계정과목 찾기
                    value = "-"
                    unit = ""
                    for stmt in statements:
                        if stmt.get('account_id', '').strip() == account_id:
                            amount = stmt.get('thstrm_amount', 0)
                            # 금액 단위 확인
                            unit = account_manager.detect_amount_unit([stmt])
                            # 안전한 금액 포맷팅 사용
                            formatted_amount = self._safe_format_amount(amount)
                            if formatted_amount != "-":
                                value = formatted_amount
                                if unit and unit != "원":
                                    value += f" {unit}"
                            break

                    ws.cell(row=row_idx, column=col, value=value)

                row_idx += 1

        # 열 너비 조정
        for column in ws.columns:
            max_length = 0
            # MergedCell 처리를 위한 예외 처리 추가
            try:
                if hasattr(column[0], 'column_letter'):
                    column_letter = column[0].column_letter
                else:
                    continue
            except:
                continue

            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if max_length > 0:
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width

    def _create_notes_sheet(self, wb: Workbook, note_data: Dict[str, Any]):
        """주석 정보 시트 생성"""
        ws = wb.create_sheet("주석정보")

        ws['A1'] = "주석 정보 상세"
        ws['A1'].font = self.title_font

        row_idx = 3
        headers = ['기업명', '연도', '감사인', '회계기준', '감가상각정책', '기타']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        row_idx = 4
        for company, years_data in note_data.items():
            for year, info in years_data.items():
                ws.cell(row=row_idx, column=1, value=company)
                ws.cell(row=row_idx, column=2, value=year)
                ws.cell(row=row_idx, column=3, value=info.get('auditor', '-'))
                ws.cell(row=row_idx, column=4, value=info.get('accounting_standard', '-'))

                # depreciation_policy가 dict인 경우 문자열로 변환
                dep_policy = info.get('depreciation_policy', '-')
                if isinstance(dep_policy, dict):
                    dep_policy = ', '.join([f"{k}: {v}" for k, v in dep_policy.items()])
                ws.cell(row=row_idx, column=5, value=dep_policy)

                ws.cell(row=row_idx, column=6, value=info.get('other', '-'))
                row_idx += 1

    def _extract_years(self, comparison_data: Dict[str, Any]) -> set:
        """데이터에서 연도 추출"""
        years = set()
        for data in comparison_data.values():
            year = data.get('year')
            if year:
                years.add(year)
        return years


# 싱글톤 인스턴스
excel_generator = ExcelGenerator()